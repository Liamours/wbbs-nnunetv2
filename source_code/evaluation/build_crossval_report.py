"""
Build crossval_comparison.xlsx from hardcoded fold DSC values.

Outputs per-class DSC for each fold and the ensemble, with paper values
side by side for both Anterior and Posterior views.

Usage:
    python build_crossval_report.py
Output saved to crossval_comparison.xlsx in the current working directory.
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

segments = [
    "skull", "cervical vert", "thoracic vert", "rib", "sternum",
    "collarbone", "scapula", "humerus", "lumbar vert", "sacrum", "pelvis", "femur"
]

paper = {
    "Anterior": {
        "skull":         (0.953, 0.02),
        "cervical vert": (0.767, 0.10),
        "thoracic vert": (0.578, 0.18),
        "rib":           (0.894, 0.04),
        "sternum":       (0.829, 0.08),
        "collarbone":    (0.700, 0.12),
        "scapula":       (0.760, 0.09),
        "humerus":       (0.859, 0.04),
        "lumbar vert":   (0.829, 0.07),
        "sacrum":        (0.825, 0.06),
        "pelvis":        (0.901, 0.03),
        "femur":         (0.879, 0.03),
        "avg":           (0.814, 0.06),
    },
    "Posterior": {
        "skull":         (0.958, 0.02),
        "cervical vert": (0.859, 0.06),
        "thoracic vert": (0.858, 0.06),
        "rib":           (0.909, 0.04),
        "sternum":       None,
        "collarbone":    (0.411, 0.22),
        "scapula":       (0.874, 0.05),
        "humerus":       (0.863, 0.06),
        "lumbar vert":   (0.847, 0.07),
        "sacrum":        (0.816, 0.07),
        "pelvis":        (0.905, 0.04),
        "femur":         (0.885, 0.06),
        "avg":           (0.835, 0.04),
    }
}

fold_data = {
    "Anterior": {
        0: [0.9526, 0.7726, 0.5726, 0.8966, 0.8386, 0.7010, 0.7623, 0.8606, 0.8367, 0.8227, 0.8983, 0.8797],
        1: [0.9520, 0.7761, 0.5651, 0.8988, 0.8405, 0.7061, 0.7636, 0.8597, 0.8363, 0.8251, 0.9006, 0.8793],
        2: [0.9542, 0.7776, 0.5750, 0.9009, 0.8432, 0.7070, 0.7602, 0.8630, 0.8408, 0.8232, 0.9000, 0.8788],
        3: [0.9523, 0.7689, 0.5639, 0.8935, 0.8320, 0.6984, 0.7596, 0.8572, 0.8336, 0.8178, 0.8973, 0.8785],
        4: [0.9516, 0.7814, 0.5849, 0.8982, 0.8414, 0.7040, 0.7643, 0.8587, 0.8417, 0.8249, 0.8988, 0.8779],
        "e": [0.9542, 0.7842, 0.5792, 0.9010, 0.8448, 0.7112, 0.7675, 0.8644, 0.8405, 0.8284, 0.9023, 0.8826],
    },
    "Posterior": {
        0: [0.9570, 0.8556, 0.8629, 0.9094, None, 0.4266, 0.8723, 0.8604, 0.8535, 0.8095, 0.9029, 0.8833],
        1: [0.9567, 0.8567, 0.8595, 0.9078, None, 0.4360, 0.8722, 0.8581, 0.8460, 0.8101, 0.9038, 0.8828],
        2: [0.9572, 0.8552, 0.8575, 0.9088, None, 0.4211, 0.8699, 0.8598, 0.8470, 0.8118, 0.9029, 0.8830],
        3: [0.9566, 0.8543, 0.8540, 0.9070, None, 0.4098, 0.8707, 0.8581, 0.8448, 0.8134, 0.9034, 0.8815],
        4: [0.9564, 0.8535, 0.8598, 0.9089, None, 0.4279, 0.8704, 0.8589, 0.8467, 0.8131, 0.9032, 0.8830],
        "e": [0.9579, 0.8587, 0.8623, 0.9113, None, 0.4343, 0.8752, 0.8631, 0.8505, 0.8156, 0.9055, 0.8859],
    }
}

thin = Side(style="thin")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR  = PatternFill("solid", start_color="BDD7EE")
HDR2 = PatternFill("solid", start_color="D9E1F2")
AVG  = PatternFill("solid", start_color="FFFF99")
NA   = PatternFill("solid", start_color="D9D9D9")
ENS  = PatternFill("solid", start_color="E2EFDA")
BOLD = Font(name="Arial", bold=True)
NORM = Font(name="Arial")
CTR  = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")


def c(cell, val, bold=False, fill=None, align="center"):
    cell.value   = val
    cell.font    = BOLD if bold else NORM
    cell.border  = BDR
    cell.alignment = CTR if align == "center" else LEFT
    if fill:
        cell.fill = fill
    if isinstance(val, float):
        cell.number_format = "0.000"


wb    = Workbook()
first = True

for view in ["Anterior", "Posterior"]:
    ws = wb.active if first else wb.create_sheet(view)
    ws.title = view
    first = False
    ws.freeze_panes = "B3"

    skip_sternum = (view == "Posterior")

    # Row 1 — group headers (merged)
    groups = [
        (1, "Segment",          1),
        (2, "Paper DSC",        2),
        (4, "Fold 0",           1),
        (5, "Fold 1",           1),
        (6, "Fold 2",           1),
        (7, "Fold 3",           1),
        (8, "Fold 4",           1),
        (9, "CV Mean \u00b1 Std", 2),
        (11, "Ensemble",        1),
    ]
    for start_col, label, span in groups:
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.font, cell.fill, cell.alignment, cell.border = BOLD, HDR, CTR, BDR
        if span > 1:
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1, end_column=start_col + span - 1)

    # Row 2 — sub-headers
    sub_headers = ["Segment", "Mean", "Std", "Fold 0", "Fold 1", "Fold 2", "Fold 3", "Fold 4", "CV Mean", "CV Std", "Ensemble"]
    for col, h in enumerate(sub_headers, 1):
        c(ws.cell(row=2, column=col), h, bold=True, fill=HDR2)

    widths = [18, 10, 8, 8, 8, 8, 8, 8, 10, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for row, seg in enumerate(segments, 3):
        is_na   = skip_sternum and seg == "sternum"
        p       = paper[view].get(seg)
        idx     = segments.index(seg)
        fvals   = [fold_data[view][f][idx] for f in range(5)]
        ens_val = fold_data[view]["e"][idx]
        fill    = NA if is_na else None

        if is_na or any(v is None for v in fvals):
            cv_mean, cv_std = None, None
        else:
            cv_mean = float(np.mean(fvals))
            cv_std  = float(np.std(fvals))

        c(ws.cell(row=row, column=1),  seg,                                   fill=fill, align="left")
        c(ws.cell(row=row, column=2),  "N/A" if is_na else (p[0] if p else "-"), fill=fill)
        c(ws.cell(row=row, column=3),  "N/A" if is_na else (p[1] if p else "-"), fill=fill)
        for fi, fv in enumerate(fvals):
            c(ws.cell(row=row, column=4 + fi), "N/A" if is_na else fv, fill=fill)
        c(ws.cell(row=row, column=9),  "N/A" if is_na else cv_mean, fill=fill)
        c(ws.cell(row=row, column=10), "N/A" if is_na else cv_std,  fill=fill)
        c(ws.cell(row=row, column=11), "N/A" if is_na else ens_val, fill=ENS)

    # Average row
    avg_row   = 15
    p_avg     = paper[view]["avg"]
    fold_avgs = []
    for f in range(5):
        vals = [v for v in fold_data[view][f] if v is not None]
        fold_avgs.append(float(np.mean(vals)))
    cv_avg_mean = float(np.mean(fold_avgs))
    cv_avg_std  = float(np.std(fold_avgs))
    ens_vals    = [v for v in fold_data[view]["e"] if v is not None]
    ens_avg     = float(np.mean(ens_vals))

    c(ws.cell(row=avg_row, column=1),  "AVERAGE",                bold=True, fill=AVG, align="left")
    c(ws.cell(row=avg_row, column=2),  p_avg[0],                 bold=True, fill=AVG)
    c(ws.cell(row=avg_row, column=3),  p_avg[1],                 bold=True, fill=AVG)
    for fi, fv in enumerate(fold_avgs):
        c(ws.cell(row=avg_row, column=4 + fi), round(fv, 4),     bold=True, fill=AVG)
    c(ws.cell(row=avg_row, column=9),  round(cv_avg_mean, 4),    bold=True, fill=AVG)
    c(ws.cell(row=avg_row, column=10), round(cv_avg_std, 4),     bold=True, fill=AVG)
    c(ws.cell(row=avg_row, column=11), round(ens_avg, 4),        bold=True, fill=ENS)

out = Path("crossval_comparison.xlsx")
wb.save(out)
print("Saved:", out)
