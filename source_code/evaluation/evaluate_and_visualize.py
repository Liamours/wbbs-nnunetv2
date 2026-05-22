"""
Evaluate nnUNetv2 predictions against ground-truth test labels.

Computes DSC and IoU per class for each fold (0-4) and ensemble,
saves an Excel report, and writes side-by-side visualizations.

Requires standard nnUNetv2 environment variables:
    nnUNet_raw   path to nnUNet_raw/

Usage:
    python evaluate_and_visualize.py --dataset 1
    python evaluate_and_visualize.py --dataset 2
    python evaluate_and_visualize.py --dataset 1 2 --report ./report
"""

import argparse
import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAW       = Path(os.environ["nnUNet_raw"])
PRED_DIRS = [f"labelsTs_pred_{i}" for i in range(5)] + ["labelsTs_pred_e"]

DATASET_META = {
    1: {"name": "Dataset001_BoneScanAnterior",  "view": "Anterior",  "prefix": "Anterior_"},
    2: {"name": "Dataset002_BoneScanPosterior", "view": "Posterior", "prefix": "Posterior_"},
}

CLASSES = {
    1: "skull", 2: "cervical vert", 3: "thoracic vert", 4: "rib",
    5: "sternum", 6: "collarbone", 7: "scapula", 8: "humerus",
    9: "lumbar vert", 10: "sacrum", 11: "pelvis", 12: "femur",
}

COLORS = [
    [0,   0,   0  ],  # 0  background
    [176, 230, 13 ],  # 1  skull
    [0,   151, 219],  # 2  cervical vertebrae
    [126, 230, 225],  # 3  thoracic vertebrae
    [166, 55,  167],  # 4  rib
    [230, 157, 180],  # 5  sternum (anterior only)
    [167, 110, 77 ],  # 6  collarbone
    [121, 0,   24 ],  # 7  scapula
    [56,  65,  184],  # 8  humerus
    [230, 218, 0  ],  # 9  lumbar vertebrae
    [230, 114, 35 ],  # 10 sacrum
    [12,  187, 62 ],  # 11 pelvis
    [230, 182, 22 ],  # 12 femur
]


def mask_to_color(mask):
    rgb = np.zeros((*mask.shape, 3), dtype=np.uint8)
    for cls_id, color in enumerate(COLORS):
        rgb[mask == cls_id] = color
    return rgb


def dice(pred, gt, cls):
    p, g = (pred == cls), (gt == cls)
    inter = (p & g).sum()
    denom = p.sum() + g.sum()
    return (2 * inter / denom) if denom > 0 else float("nan")


def iou(pred, gt, cls):
    p, g = (pred == cls), (gt == cls)
    inter = (p & g).sum()
    union = (p | g).sum()
    return (inter / union) if union > 0 else float("nan")


def pred_stem_to_gt_name(stem, prefix):
    return (stem.replace(prefix, "").lstrip("0") or "0") + ".png"


def evaluate(pred_dir, gt_dir, dataset_id):
    prefix       = DATASET_META[dataset_id]["prefix"]
    skip_classes = {0} | ({5} if dataset_id == 2 else set())
    all_dice     = {c: [] for c in range(1, 13)}
    all_iou      = {c: [] for c in range(1, 13)}

    pred_files = sorted(pred_dir.glob("*.png"))
    matched = 0
    for pf in pred_files:
        gt_path = gt_dir / pred_stem_to_gt_name(pf.stem, prefix)
        if not gt_path.exists():
            continue
        matched += 1
        pred_arr = np.array(Image.open(pf))
        gt_arr   = np.array(Image.open(gt_path))
        for c in range(1, 13):
            if c in skip_classes:
                continue
            all_dice[c].append(dice(pred_arr, gt_arr, c))
            all_iou[c].append(iou(pred_arr, gt_arr, c))

    print(f"\n  [{pred_dir.name}]  {matched} images")
    print(f"  {'Class':<20} {'DSC mean':>10} {'DSC std':>10} {'IoU mean':>10} {'IoU std':>10}")
    print("  " + "-" * 62)

    results = {}
    row_dsc, row_iou = [], []
    for c in range(1, 13):
        if c in skip_classes:
            print(f"  {CLASSES[c]:<20} {'N/A':>10}")
            results[c] = None
            continue
        d_vals = [v for v in all_dice[c] if not np.isnan(v)]
        i_vals = [v for v in all_iou[c]  if not np.isnan(v)]
        if not d_vals:
            results[c] = None
            continue
        d_mean, d_std = np.mean(d_vals), np.std(d_vals)
        i_mean, i_std = np.mean(i_vals), np.std(i_vals)
        row_dsc.append(d_mean)
        row_iou.append(i_mean)
        results[c] = (d_mean, d_std, i_mean, i_std)
        print(f"  {CLASSES[c]:<20} {d_mean:>10.3f} {d_std:>10.3f} {i_mean:>10.3f} {i_std:>10.3f}")

    avg_dsc, avg_iou = np.mean(row_dsc), np.mean(row_iou)
    print("  " + "-" * 62)
    print(f"  {'AVERAGE':<20} {avg_dsc:>10.3f} {'':>10} {avg_iou:>10.3f}")
    results["avg_dsc"] = avg_dsc
    results["avg_iou"] = avg_iou
    return results


def visualize_examples(pred_dir, gt_dir, img_dir, prefix, n=5):
    out_dir = pred_dir.parent / f"{pred_dir.name}_viz"
    out_dir.mkdir(parents=True, exist_ok=True)
    for pf in sorted(pred_dir.glob("*.png"))[:n]:
        gt_path  = gt_dir / pred_stem_to_gt_name(pf.stem, prefix)
        img_path = img_dir / (pf.stem + "_0000.png")
        if not gt_path.exists() or not img_path.exists():
            continue
        img      = np.array(Image.open(img_path).convert("L"))
        gt_arr   = np.array(Image.open(gt_path))
        pred_arr = np.array(Image.open(pf))
        fig, axes = plt.subplots(1, 3, figsize=(15, 5))
        axes[0].imshow(img, cmap="gray");        axes[0].set_title("Input Image");  axes[0].axis("off")
        axes[1].imshow(mask_to_color(gt_arr));   axes[1].set_title("Ground Truth"); axes[1].axis("off")
        axes[2].imshow(mask_to_color(pred_arr)); axes[2].set_title("Prediction");   axes[2].axis("off")
        patches = [mpatches.Patch(color=[ch/255 for ch in COLORS[i]], label=CLASSES[i])
                   for i in range(1, 13)]
        fig.legend(handles=patches, loc="lower center", ncol=6, fontsize=7,
                   bbox_to_anchor=(0.5, -0.05))
        plt.suptitle(f"{pred_dir.name} | {pf.stem}")
        plt.tight_layout()
        plt.savefig(out_dir / f"{pf.stem}_comparison.png", bbox_inches="tight", dpi=100)
        plt.close()
    print(f"  Visualizations -> {out_dir}")


def save_report(all_summary, report_dir: Path):
    REPORT = report_dir
    REPORT.mkdir(parents=True, exist_ok=True)
    wb   = Workbook()
    thin = Side(style="thin")
    BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR  = PatternFill("solid", start_color="BDD7EE")
    AVG  = PatternFill("solid", start_color="FFFF99")
    NA   = PatternFill("solid", start_color="D9D9D9")
    BOLD = Font(name="Arial", bold=True)
    NORM = Font(name="Arial")
    CTR  = Alignment(horizontal="center")
    LEFT = Alignment(horizontal="left")

    first = True
    for (dataset_id, pred_name), results in all_summary.items():
        view    = DATASET_META[dataset_id]["view"]
        ds_name = DATASET_META[dataset_id]["name"]
        skip    = {5} if dataset_id == 2 else set()
        title   = f"{view}_{pred_name}"

        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        first = False
        ws.freeze_panes = "A2"

        headers   = ["Segment", "DSC Mean", "DSC Std", "IoU Mean", "IoU Std"]
        col_widths = [20, 12, 12, 12, 12]
        for c, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = BOLD, HDR, CTR, BDR
            ws.column_dimensions[get_column_letter(c)].width = w

        for r, cls_id in enumerate(range(1, 13), 2):
            is_na = cls_id in skip
            row_data = results.get(cls_id)
            vals = [CLASSES[cls_id],
                    "N/A" if is_na else (round(row_data[0], 4) if row_data else "-"),
                    "N/A" if is_na else (round(row_data[1], 4) if row_data else "-"),
                    "N/A" if is_na else (round(row_data[2], 4) if row_data else "-"),
                    "N/A" if is_na else (round(row_data[3], 4) if row_data else "-")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font, cell.border = NORM, BDR
                cell.alignment = LEFT if c == 1 else CTR
                if is_na:
                    cell.fill = NA
                if isinstance(v, float):
                    cell.number_format = "0.000"

        avg_row = 14
        avg_vals = ["AVERAGE", round(results["avg_dsc"], 4), "-",
                    round(results["avg_iou"], 4), "-"]
        for c, v in enumerate(avg_vals, 1):
            cell = ws.cell(row=avg_row, column=c, value=v)
            cell.font, cell.fill, cell.border = BOLD, AVG, BDR
            cell.alignment = LEFT if c == 1 else CTR
            if isinstance(v, float):
                cell.number_format = "0.000"

    summary_ws = wb.create_sheet("Summary")
    summary_ws.freeze_panes = "A2"
    s_headers = ["View", "Fold / Ensemble", "Avg DSC", "Avg IoU"]
    s_widths  = [12, 20, 12, 12]
    for c, (h, w) in enumerate(zip(s_headers, s_widths), 1):
        cell = summary_ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = BOLD, HDR, CTR, BDR
        summary_ws.column_dimensions[get_column_letter(c)].width = w

    for r, ((dataset_id, pred_name), results) in enumerate(all_summary.items(), 2):
        view = DATASET_META[dataset_id]["view"]
        for c, v in enumerate([view, pred_name,
                                round(results["avg_dsc"], 4),
                                round(results["avg_iou"], 4)], 1):
            cell = summary_ws.cell(row=r, column=c, value=v)
            cell.font, cell.border = NORM, BDR
            cell.alignment = LEFT if c <= 2 else CTR
            if isinstance(v, float):
                cell.number_format = "0.000"

    out_path = REPORT / "evaluation_results.xlsx"
    wb.save(out_path)
    print(f"\nReport saved -> {out_path}")


def process_dataset(dataset_id, all_summary):
    meta    = DATASET_META[dataset_id]
    ds_name = meta["name"]
    view    = meta["view"]
    prefix  = meta["prefix"]
    gt_dir  = RAW / ds_name / "labelsTs"
    img_dir = RAW / ds_name / "imagesTs"

    if not gt_dir.exists():
        print(f"GT not found: {gt_dir}")
        return

    print(f"\n{'='*66}\n  {view}  —  {ds_name}\n{'='*66}")

    for pred_name in PRED_DIRS:
        pred_dir = RAW / ds_name / pred_name
        if not pred_dir.exists():
            print(f"\n  [{pred_name}] NOT FOUND — skipping")
            continue
        results = evaluate(pred_dir, gt_dir, dataset_id)
        all_summary[(dataset_id, pred_name)] = results
        visualize_examples(pred_dir, gt_dir, img_dir, prefix, n=5)

    print(f"\n{'='*66}\n  SUMMARY — {view}")
    print(f"  {'Pred Dir':<22} {'Avg DSC':>10} {'Avg IoU':>10}\n  " + "-" * 44)
    for (ds_id, pred_name), res in all_summary.items():
        if ds_id != dataset_id:
            continue
        print(f"  {pred_name:<22} {res['avg_dsc']:>10.3f} {res['avg_iou']:>10.3f}")
    print(f"{'='*66}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", type=int, nargs="+", choices=[1, 2], default=[1, 2])
    parser.add_argument("--report", type=Path, default=Path("report"),
                        help="Directory for Excel report output (default: ./report)")
    args = parser.parse_args()

    all_summary = {}
    for ds_id in args.dataset:
        process_dataset(ds_id, all_summary)

    save_report(all_summary, args.report)


if __name__ == "__main__":
    main()
